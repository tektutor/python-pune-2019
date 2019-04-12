#!/usr/bin/python3

import openpyxl

def readFromExcel(fileName):

    workBook = openpyxl.load_workbook(fileName)
    sheet = workBook.get_sheet_by_name("Sheet1")

    for row in range(1, sheet.max_row+1):
        for col in range(1, sheet.max_column+1):
            cell = sheet.cell(row,col)
            print ( cell.value )

    workBook.close()


def writeToExcel(fileName):

    workBook = openpyxl.load_workbook(fileName)
    sheet = workBook.get_sheet_by_name("Sheet1")
    heading4 = sheet.cell(1,4)
    heading4.value = "New Column"

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row,4)
        cell.value = 100 * row

    workBook.save(fileName)
    workBook.close()

def copyWorkBook(sourceFileName, destFileName):











writeToExcel("/home/jegan/Documents/Training.xlsx")
readFromExcel("/home/jegan/Documents/Training.xlsx")
